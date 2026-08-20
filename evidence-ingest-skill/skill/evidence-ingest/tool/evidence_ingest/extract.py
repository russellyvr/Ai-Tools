"""Deterministic extraction: bytes in the vault -> canonical ExtractionBundle.

Parsers are deterministic (stdlib plus pinned pure-Python readers for OOXML):
  * ``.eml`` — ``email.parser.BytesParser`` (compat32 verbatim headers +
    getaddresses recipients + best-effort text body walk).
  * ``.txt``/``.md``/``.csv`` — UTF-8 with ``errors="replace"``; any
    replacement is recorded as a parse warning (bytes in the vault stay
    authoritative). CSV additionally records rows/cols shape metadata.
  * ``.html``/``.htm`` — stdlib ``html.parser`` visible-text extraction:
    scripts never executed, links/resources never fetched;
    ``<script>``/``<style>``/``<template>`` blocks and comments are excluded
    from the text and counted as parse warnings; ``<title>`` is captured as
    the bundle subject.
  * ``.xlsx`` — openpyxl read-only, ``data_only=False``: formulas stay as
    source text, never evaluated; external links never followed; one Page
    per worksheet as tab-separated lines.
  * ``.docx`` — python-docx, paragraphs and tables in body order; macros,
    OLE, and external relationships never executed or followed.
  * ``.doc``/``.xls`` (legacy OLE) — no deterministic parser: preserved
    verbatim with a ``needs_conversion`` warning (LibreOffice headless lane).
  * PDF / images — OCR lane: if ``ocr-raw/<sha>.json`` exists the stored
    Azure Read / Google Document AI result is parsed from that artifact;
    otherwise the bundle is opaque (``channel: none``) with an
    ``ocr_unavailable`` warning.

Every bundle is written as canonical JSON; ``extraction_sha256`` used by the
chunker is the sha256 of the bundle file bytes.
"""
from __future__ import annotations

import json
from email import policy
from email.header import decode_header, make_header
from email.parser import BytesParser
from email.utils import getaddresses
from pathlib import Path

from evidence_ingest import TOOL_VERSION
from evidence_ingest.custody import CustodyLog
from evidence_ingest.hashing import canonical_json
from evidence_ingest.scan import load_records
from evidence_ingest.schemas import (
    EmailRecipient, EvidenceRecord, ExtractionBundle, Page, TableMeta)


def _decoded(value) -> str:
    if not value:
        return ""
    try:
        return str(make_header(decode_header(str(value)))).replace("\r", " ").replace("\n", " ").strip()
    except Exception:
        return str(value).replace("\r", " ").replace("\n", " ").strip()


def extract_eml(data: bytes) -> ExtractionBundle | dict:
    """Parse an RFC5322 message deterministically; verbatim header capture."""
    warnings: list[str] = []
    msg = BytesParser(policy=policy.compat32).parsebytes(data)

    raw_headers: dict[str, list[str]] = {}
    for key in msg.keys():
        raw_headers.setdefault(key, [])
    for key in raw_headers:
        raw_headers[key] = [str(v) for v in msg.get_all(key) or []]

    recipients: list[EmailRecipient] = []
    for role, field in (("from", "From"), ("sender", "Sender"),
                        ("reply_to", "Reply-To"), ("to", "To"),
                        ("cc", "Cc"), ("bcc", "Bcc")):
        values = msg.get_all(field)
        if not values:
            continue
        for display, addr in getaddresses([str(v) for v in values]):
            if addr:
                recipients.append(EmailRecipient(
                    role=role, display_name=_decoded(display) or None,
                    address=addr.strip().lower(), raw=f"{display} <{addr}>".strip()))

    body_parts: list[str] = []
    if msg.is_multipart():
        for part in msg.walk():
            if part.get_content_maintype() != "text":
                continue
            try:
                payload = part.get_payload(decode=True)
            except Exception as e:
                warnings.append(f"payload decode failed: {e}")
                continue
            if payload is None:
                continue
            charset = part.get_content_charset() or "utf-8"
            try:
                text = payload.decode(charset, errors="replace")
            except LookupError:
                text = payload.decode("utf-8", errors="replace")
                warnings.append(f"unknown charset {charset!r}; decoded as utf-8/replace")
            if part.get_content_subtype() == "plain":
                body_parts.append(text)
    else:
        payload = msg.get_payload(decode=True)
        if payload is None:
            payload = str(msg.get_payload()).encode("utf-8", errors="replace")
        charset = msg.get_content_charset() or "utf-8"
        try:
            text = payload.decode(charset, errors="replace")
        except LookupError:
            text = payload.decode("utf-8", errors="replace")
            warnings.append(f"unknown charset {charset!r}; decoded as utf-8/replace")
        body_parts.append(text)

    body = "\n".join(body_parts)
    if "\ufffd" in body:
        warnings.append("body contains U+FFFD replacement characters (lossy decode)")

    return dict(
        channel="native-text",
        pages=[Page(page_number=None, text=body)],
        parse_warnings=warnings,
        subject=_decoded(msg.get("Subject")) or None,
        date_raw=(msg.get("Date") or "").strip() or None,
        message_id=(msg.get("Message-ID") or "").strip() or None,
        recipients=recipients,
        raw_headers=raw_headers,
    )


def extract_txt(data: bytes) -> dict:
    warnings: list[str] = []
    try:
        text = data.decode("utf-8")
    except UnicodeDecodeError:
        text = data.decode("utf-8", errors="replace")
        warnings.append("invalid UTF-8 bytes replaced with U+FFFD (source bytes remain authoritative in vault)")
    return dict(channel="native-text",
                pages=[Page(page_number=None, text=text)],
                parse_warnings=warnings)


def extract_csv(data: bytes) -> dict:
    """CSV: the decoded text is preserved verbatim as the page text (bytes in
    the vault stay authoritative); rows/cols are counted with the stdlib csv
    reader (excel dialect) purely as shape metadata. Formula-leading cells
    are never evaluated — they remain inert text."""
    import csv as _csv
    import io

    warnings: list[str] = []
    try:
        text = data.decode("utf-8")
    except UnicodeDecodeError:
        text = data.decode("utf-8", errors="replace")
        warnings.append("invalid UTF-8 bytes replaced with U+FFFD (source bytes remain authoritative in vault)")
    rows = 0
    cols = 0
    try:
        for row in _csv.reader(io.StringIO(text)):
            rows += 1
            if len(row) > cols:
                cols = len(row)
    except _csv.Error as e:
        warnings.append(f"csv shape scan failed: {e}")
    return dict(channel="native-text",
                pages=[Page(page_number=None, text=text)],
                parse_warnings=warnings,
                tables_meta=[TableMeta(name="csv", page_number=1,
                                       rows=rows, cols=cols)])


_HTML_SKIP_TAGS = ("script", "style", "template")
_HTML_BLOCK_TAGS = (
    "address", "article", "aside", "blockquote", "body", "caption", "dd",
    "details", "dialog", "div", "dl", "dt", "fieldset", "figcaption",
    "figure", "footer", "form", "h1", "h2", "h3", "h4", "h5", "h6", "head",
    "header", "hr", "html", "li", "main", "nav", "ol", "option", "p", "pre",
    "section", "summary", "table", "tbody", "tfoot", "thead", "tr", "ul")


class _VisibleTextHTMLParser:
    """Deterministic stdlib-only visible-text extraction.

    Scripts are never executed; styles never applied; links and external
    resources never fetched. ``<script>``/``<style>``/``<template>``
    contents and comments are excluded from the emitted text (the counts
    are surfaced as parse warnings by ``extract_html``; the verbatim bytes
    stay authoritative in the vault). ``<title>`` text is captured
    separately for bundle metadata.
    """

    def __init__(self) -> None:
        from html.parser import HTMLParser

        outer = self

        class _P(HTMLParser):
            def handle_starttag(self, tag, attrs):
                outer._start(tag)

            def handle_startendtag(self, tag, attrs):
                outer._start(tag)
                outer._end(tag)

            def handle_endtag(self, tag):
                outer._end(tag)

            def handle_data(self, data):
                outer._data(data)

            def handle_comment(self, data):
                outer.comment_count += 1

        self._parser = _P(convert_charrefs=True)
        self._pieces: list[str] = []
        self._skip_depth = 0
        self._in_title = False
        self._title_pieces: list[str] = []
        self.skipped_blocks = 0
        self.comment_count = 0

    def _start(self, tag: str) -> None:
        if tag in _HTML_SKIP_TAGS:
            self._skip_depth += 1
            self.skipped_blocks += 1
        elif tag == "title":
            self._in_title = True
        elif tag == "br":
            self._pieces.append("\n")
        elif tag in ("td", "th"):
            self._pieces.append("\t")
        elif tag in _HTML_BLOCK_TAGS:
            self._pieces.append("\n")

    def _end(self, tag: str) -> None:
        if tag in _HTML_SKIP_TAGS:
            if self._skip_depth > 0:
                self._skip_depth -= 1
        elif tag == "title":
            self._in_title = False
        elif tag in _HTML_BLOCK_TAGS:
            self._pieces.append("\n")

    def _data(self, data: str) -> None:
        if self._skip_depth > 0:
            return
        if self._in_title:
            self._title_pieces.append(data)
            return
        self._pieces.append(data)

    def feed(self, text: str) -> None:
        self._parser.feed(text)
        self._parser.close()

    @property
    def title(self) -> str:
        return " ".join("".join(self._title_pieces).split())

    def text(self) -> str:
        lines = [ln.rstrip() for ln in "".join(self._pieces).split("\n")]
        out: list[str] = []
        for ln in lines:
            if ln == "" and (not out or out[-1] == ""):
                continue
            out.append(ln)
        while out and out[-1] == "":
            out.pop()
        return "\n".join(out)


def extract_html(data: bytes) -> dict:
    """HTML via stdlib ``html.parser`` only (no third-party libraries):
    visible text with block-level newlines, entities decoded. Scripts are
    never executed, external resources never fetched. ``<script>``/
    ``<style>``/``<template>`` contents and comments are excluded from the
    page text and counted in a parse warning so hidden content is never
    silently invisible — the vault bytes remain authoritative."""
    warnings: list[str] = []
    try:
        markup = data.decode("utf-8")
    except UnicodeDecodeError:
        markup = data.decode("utf-8", errors="replace")
        warnings.append("invalid UTF-8 bytes replaced with U+FFFD (source bytes remain authoritative in vault)")
    parser = _VisibleTextHTMLParser()
    try:
        parser.feed(markup)
        text = parser.text()
        title = parser.title or None
    except Exception as e:  # fail-open to verbatim markup, never crash the run
        warnings.append(f"html parse failed ({type(e).__name__}: {e}); raw markup preserved as page text")
        return dict(channel="native-text",
                    pages=[Page(page_number=None, text=markup)],
                    parse_warnings=warnings)
    if parser.skipped_blocks:
        warnings.append(
            f"html: {parser.skipped_blocks} script/style/template block(s) "
            "excluded from extracted text (source bytes remain authoritative in vault)")
    if parser.comment_count:
        warnings.append(
            f"html: {parser.comment_count} comment(s) excluded from extracted "
            "text (source bytes remain authoritative in vault)")
    return dict(channel="native-text",
                pages=[Page(page_number=None, text=text)],
                parse_warnings=warnings,
                subject=title)


def _cell_str(v) -> str:
    """Canonical, deterministic cell rendering. Formulas stay as their
    source text; nothing is ever computed. Tabs/newlines inside a cell are
    flattened to spaces so the TSV serialization stays unambiguous."""
    if v is None:
        return ""
    if v is True:
        return "TRUE"
    if v is False:
        return "FALSE"
    if hasattr(v, "isoformat"):
        return v.isoformat()
    return str(v).replace("\t", " ").replace("\r", " ").replace("\n", " ")


def extract_xlsx(data: bytes) -> dict:
    """XLSX via openpyxl read-only: one Page per worksheet, cells serialized
    as tab-separated lines in stored order. Formulas are captured as their
    source text (``data_only=False``) — never evaluated; external links,
    macros, and defined names are never followed or executed."""
    import io

    warnings: list[str] = []
    try:
        from openpyxl import load_workbook
    except ImportError:
        return dict(channel="none", pages=[],
                    parse_warnings=["openpyxl not installed; xlsx preserved verbatim in vault"])
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=False,
                           keep_links=False)
    except Exception as e:
        return dict(channel="none", pages=[],
                    parse_warnings=[f"xlsx unparseable: {type(e).__name__}: {e}"])
    pages: list[Page] = []
    metas: list[TableMeta] = []
    try:
        for i, ws in enumerate(wb.worksheets, start=1):
            lines: list[str] = []
            n_rows = 0
            n_cols = 0
            for row in ws.iter_rows(values_only=True):
                n_rows += 1
                if len(row) > n_cols:
                    n_cols = len(row)
                lines.append("\t".join(_cell_str(c) for c in row))
            pages.append(Page(page_number=i, text="\n".join(lines)))
            metas.append(TableMeta(name=str(ws.title), page_number=i,
                                   rows=n_rows, cols=n_cols))
    finally:
        wb.close()
    return dict(channel="native-text", pages=pages, parse_warnings=warnings,
                tables_meta=metas)


def extract_docx(data: bytes) -> dict:
    """DOCX via python-docx: paragraphs and tables in document body order.
    Tables serialize as tab-separated lines. Macros, OLE objects, and
    external relationships are never executed or followed."""
    import io

    try:
        import docx as _docx
        from docx.table import Table as _Table
        from docx.text.paragraph import Paragraph as _Paragraph
    except ImportError:
        return dict(channel="none", pages=[],
                    parse_warnings=["python-docx not installed; docx preserved verbatim in vault"])
    try:
        doc = _docx.Document(io.BytesIO(data))
    except Exception as e:
        return dict(channel="none", pages=[],
                    parse_warnings=[f"docx unparseable: {type(e).__name__}: {e}"])

    blocks: list[str] = []
    body = doc.element.body
    for child in body.iterchildren():
        tag = child.tag.rsplit("}", 1)[-1]
        if tag == "p":
            blocks.append(_Paragraph(child, doc).text)
        elif tag == "tbl":
            table = _Table(child, doc)
            rows = []
            for row in table.rows:
                rows.append("\t".join(
                    _cell_str(cell.text) for cell in row.cells))
            blocks.append("\n".join(rows))
    return dict(channel="native-text",
                pages=[Page(page_number=None, text="\n".join(blocks))],
                parse_warnings=[])


def parse_azure_read_json(raw: bytes) -> list[Page]:
    """Extract per-page text from a stored Azure Read 3.2 result, verbatim
    line order, LF-joined."""
    doc = json.loads(raw)
    pages: list[Page] = []
    results = (doc.get("analyzeResult") or {}).get("readResults") or []
    for pr in results:
        lines = [ln.get("text", "") for ln in pr.get("lines", [])]
        pages.append(Page(page_number=int(pr.get("page", len(pages) + 1)),
                          text="\n".join(lines)))
    return pages


def parse_google_docai_json(raw: bytes) -> list[Page]:
    """Extract per-page text from a stored Google Document AI ``:process``
    response.

    The response carries ``document.text`` (full text) and
    ``document.pages[]``; each page's text is re-derived from
    ``pages[].layout.textAnchor.textSegments`` — (startIndex, endIndex)
    offsets into ``document.text``, where an absent index means 0. Purely
    positional, deterministic, no interpretation.
    """
    doc = json.loads(raw).get("document") or {}
    return _pages_from_docai_document(doc)


def _pages_from_docai_document(doc: dict, page_offset: int = 0) -> list[Page]:
    """Positional page-text re-derivation from one Document AI ``document``
    object; ``page_offset`` renumbers pages of an assembled part back to
    their position in the original file."""
    full_text = doc.get("text", "")
    pages: list[Page] = []
    for i, pg in enumerate(doc.get("pages") or [], start=1):
        segments = (((pg.get("layout") or {}).get("textAnchor") or {})
                    .get("textSegments") or [])
        parts = []
        for seg in segments:
            start = int(seg.get("startIndex", 0))
            end = int(seg.get("endIndex", 0))
            parts.append(full_text[start:end])
        pages.append(Page(page_number=int(pg.get("pageNumber", i)) + page_offset,
                          text="".join(parts)))
    if not pages and full_text:
        pages.append(Page(page_number=1 + page_offset, text=full_text))
    return pages


def parse_google_docai_assembled(raw: bytes) -> list[Page]:
    """Parse an assembled multi-part Document AI artifact (an oversized PDF
    deterministically split into sync-sized slices at OCR time). Each part
    embeds its verbatim ``:process`` response plus the 1-based source page
    range; page numbers are re-derived against the original document."""
    env = json.loads(raw)
    pages: list[Page] = []
    for part in env.get("parts") or []:
        doc = (part.get("response") or {}).get("document") or {}
        pages.extend(_pages_from_docai_document(
            doc, page_offset=int(part.get("page_start", 1)) - 1))
    return pages


def parse_ocr_raw(raw: bytes) -> tuple[str, list[Page]]:
    """Route a stored OCR artifact to its parser by response shape.

    Azure Read responses carry ``analyzeResult``; Document AI responses
    carry ``document``. Returns (channel, pages).
    """
    doc = json.loads(raw)
    if "analyzeResult" in doc:
        return "azure-read-3.2", parse_azure_read_json(raw)
    if doc.get("assembled_from_parts"):
        return "google-docai-v1", parse_google_docai_assembled(raw)
    if "document" in doc:
        return "google-docai-v1", parse_google_docai_json(raw)
    raise ValueError("unrecognized OCR response shape")


def _bundle_for(rec: EvidenceRecord, work: Path) -> ExtractionBundle:
    data = (work / rec.vault_relpath).read_bytes()
    common = dict(source_sha256=rec.sha256, media_type=rec.media_type,
                  tool_version=TOOL_VERSION)
    if rec.media_type == "eml":
        return ExtractionBundle(**common, **extract_eml(data))
    if rec.media_type == "txt":
        return ExtractionBundle(**common, **extract_txt(data))
    if rec.media_type == "csv":
        return ExtractionBundle(**common, **extract_csv(data))
    if rec.media_type == "html":
        return ExtractionBundle(**common, **extract_html(data))
    if rec.media_type == "xlsx":
        return ExtractionBundle(**common, **extract_xlsx(data))
    if rec.media_type == "docx":
        return ExtractionBundle(**common, **extract_docx(data))
    if rec.media_type in ("doc", "xls"):
        # Legacy OLE formats: no deterministic pure-Python parser. Bytes are
        # preserved verbatim in the vault; convert to docx/xlsx with
        # LibreOffice headless (offline, macros disabled) and re-ingest.
        return ExtractionBundle(
            **common, channel="none", pages=[],
            parse_warnings=["needs_conversion: convert to OOXML via "
                            "'soffice --headless --convert-to' and re-ingest"])
    if rec.media_type in ("pdf", "png", "jpeg", "tiff", "bmp"):
        raw_path = work / "ocr-raw" / f"{rec.sha256}.json"
        if raw_path.exists():
            raw = raw_path.read_bytes()
            from evidence_ingest.hashing import sha256_of_bytes
            try:
                channel, pages = parse_ocr_raw(raw)
                return ExtractionBundle(
                    **common, channel=channel, pages=pages,
                    parse_warnings=[], ocr_raw_sha256=sha256_of_bytes(raw))
            except (json.JSONDecodeError, ValueError, TypeError) as e:
                return ExtractionBundle(
                    **common, channel="none", pages=[],
                    parse_warnings=[f"ocr_result_unparseable: {e}"],
                    ocr_raw_sha256=sha256_of_bytes(raw))
        return ExtractionBundle(**common, channel="none", pages=[],
                                parse_warnings=["ocr_unavailable"])
    # opaque/other: no text channel; bytes preserved in vault only
    return ExtractionBundle(**common, channel="none", pages=[],
                            parse_warnings=["no deterministic parser for media type"])


def run_extract(work: Path, custody: CustodyLog) -> int:
    """Write ``extracted/<sha>.bundle.json`` for every accepted record."""
    work = Path(work)
    records = load_records(work)
    out_dir = work / "extracted"
    out_dir.mkdir(exist_ok=True)
    n = 0
    for sha in sorted(records):
        bundle = _bundle_for(records[sha], work)
        path = out_dir / f"{sha}.bundle.json"
        path.write_text(canonical_json(bundle.model_dump()) + "\n",
                        encoding="utf-8", newline="\n")
        n += 1
        custody.append("extraction_written", {
            "sha256": sha, "channel": bundle.channel,
            "parse_warnings": len(bundle.parse_warnings)})
    custody.append("extract_completed", {"bundles": n})
    return n


def load_bundle(work: Path, sha: str) -> tuple[ExtractionBundle, str]:
    """Load a bundle and its extraction_sha256 (hash of the file bytes)."""
    from evidence_ingest.hashing import sha256_of_bytes
    path = Path(work) / "extracted" / f"{sha}.bundle.json"
    raw = path.read_bytes()
    bundle = ExtractionBundle.model_validate(json.loads(raw))
    return bundle, sha256_of_bytes(raw)
