"""contortionist — schema- and resource-abuse agent.

Traps: extra JSON fields and wrong types injected into pipeline artifacts
(pydantic ``extra="forbid"`` must reject), CSV-injection filenames
(``=cmd()`` cells must be neutralized in the index), control characters and
JSONL-smuggling content (chunks.jsonl must stay one-object-per-line and
round-trip), invalid UTF-8 (accepted with a recorded warning, vault bytes
verbatim), and oversized files (RESOURCE_* drop with a reconciling ledger).
"""
from __future__ import annotations

import csv
import json

from evidence_ingest.hashing import sha256_of_bytes
from evidence_ingest.report import build_reports
from evidence_ingest.schemas import REASON_MAGIC_MISMATCH, REASON_TOO_LARGE


def run(ctx) -> None:
    inp, work = ctx.dirs("main")

    oversized = b"A" * 65536
    (inp / "oversized.txt").write_bytes(oversized)

    csv_name = "=cmd()@SUM(1+9).txt"
    csv_body = b"=HYPERLINK(\"http://evil.example\",\"click\")\n+cmd|whoami\n"
    (inp / csv_name).write_bytes(csv_body)

    smuggle = (
        b"line one\n"
        b"{\"fake\": \"jsonl record\", \"event_hash\": \"0\"}\n"
        b"\x01\x02 control chars \x1f embedded\r\n"
        b"tab\tseparated\tsmuggle\n"
    )
    (inp / "smuggle.txt").write_bytes(smuggle)

    bad_utf8 = b"prefix \xff\xfe\xfa raw bytes suffix\n"
    (inp / "badutf8.txt").write_bytes(bad_utf8)

    control = b"Plain compliant document body for the control lane.\n" * 3
    (inp / "control.txt").write_bytes(control)

    # xlsx with a hostile formula cell: must ingest with the formula captured
    # as inert source text, never evaluated, no link followed
    import io as _io

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "attack"
    ws["A1"] = '=HYPERLINK("http://evil.example/x","click")'
    ws["A2"] = "legit value"
    buf = _io.BytesIO()
    wb.save(buf)
    formula_xlsx = buf.getvalue()
    (inp / "formulas.xlsx").write_bytes(formula_xlsx)

    # csv whose CELLS are formula-shaped: text must survive verbatim
    formula_csv = b'=cmd|calc!A0,plain\n@SUM(1+1),other\n'
    (inp / "formulas.csv").write_bytes(formula_csv)

    # html extension over PDF magic bytes: must drop as a magic mismatch
    masq_html = b"%PDF-1.7 pretending to be markup\n%%EOF"
    (inp / "masq.html").write_bytes(masq_html)

    # malformed html: unclosed tags, bogus entities, nested table soup —
    # must be accepted and extracted deterministically, never crash
    malformed_html = (b"<div><p<b>unclosed &notanentity; <table><tr><td nested"
                      b"<script>var x=1;<style>oops</div> trailing visible text")
    (inp / "malformed.html").write_bytes(malformed_html)

    # cap sized so the oversized trap fires while real xlsx fixtures pass
    ctx.pipeline(inp, work, max_file_bytes=32768)

    records = json.loads((work / "ledger" / "records.json").read_text(encoding="utf-8"))
    drops = {}
    with open(work / "ledger" / "drops.jsonl", encoding="utf-8") as f:
        for line in f:
            if line.strip():
                d = json.loads(line)
                drops[d["relpath"]] = d["reason_code"]

    # Trap 1: oversized -> RESOURCE_TOO_LARGE and ledger still reconciles
    ok = drops.get("oversized.txt") == REASON_TOO_LARGE
    ctx.trap("oversized-capped-with-resource-code", ok, f"got {drops.get('oversized.txt')}")
    ok, detail = ctx.validate_ok(inp, work)
    ctx.trap("ledger-reconciles-after-drop", ok, detail)

    # Trap 2: CSV-injection filename neutralized in _index.csv
    build_reports(work, work)
    with open(work / "_index.csv", encoding="utf-8", newline="") as f:
        rows = list(csv.reader(f))
    sha = sha256_of_bytes(csv_body)
    cells = [row[4] for row in rows[1:] if row[0] == sha]
    ok = bool(cells) and all(c.startswith("'") for c in cells) and sha in records
    ctx.trap("csv-injection-cell-neutralized", ok,
             "formula-leading filename reached the index unescaped")

    # Trap 3: control chars / JSONL smuggling — chunks.jsonl stays valid
    smuggle_sha = sha256_of_bytes(smuggle)
    ok = False
    detail = "smuggle content missing from chunks"
    with open(work / "rag" / "chunks.jsonl", encoding="utf-8") as f:
        lines = f.read().splitlines()
    try:
        parsed = [json.loads(ln) for ln in lines if ln]
        rebuilt = "".join(c["text"] for c in parsed
                          if c["source_sha256"] == smuggle_sha and c["char_start"] == 0)
        # canonical JSON escapes every control char; the fake JSONL record is
        # inert payload inside a string, and the text round-trips exactly
        ok = rebuilt.startswith("line one\n{\"fake\"") and "\x01\x02" in rebuilt
        detail = "" if ok else "control characters or smuggled line altered chunk text"
    except json.JSONDecodeError as e:
        detail = f"chunks.jsonl broken by smuggled content: {e}"
    ctx.trap("jsonl-smuggling-inert", ok, detail)

    # Trap 4: invalid UTF-8 accepted verbatim with recorded warning
    bad_sha = sha256_of_bytes(bad_utf8)
    bundle_path = work / "extracted" / f"{bad_sha}.bundle.json"
    ok = False
    detail = "bundle missing"
    if bad_sha in records and bundle_path.exists():
        bundle = json.loads(bundle_path.read_text(encoding="utf-8"))
        vault_ok = (work / records[bad_sha]["vault_relpath"]).read_bytes() == bad_utf8
        ok = vault_ok and any("invalid UTF-8" in w for w in bundle["parse_warnings"])
        detail = "vault bytes altered or replacement warning not recorded"
    ctx.trap("invalid-utf8-warned-not-silent", ok, detail)

    # Trap 5: extra JSON field injected into a bundle -> schema rejection
    target = work / "extracted" / f"{bad_sha}.bundle.json"
    original = target.read_bytes()
    doc = json.loads(original)
    doc["smuggled_field"] = {"nested": True}
    target.write_text(json.dumps(doc, sort_keys=True), encoding="utf-8")
    ok, _ = ctx.validate_ok(inp, work)
    ctx.trap("extra-json-field-rejected", not ok,
             "pydantic extra=forbid failed to reject unknown field")
    target.write_bytes(original)

    # Trap 6: wrong type injected into chunks.jsonl -> schema rejection
    chunks_path = work / "rag" / "chunks.jsonl"
    original = chunks_path.read_bytes()
    lines = original.decode("utf-8").splitlines()
    doc = json.loads(lines[0])
    doc["char_start"] = "zero"  # wrong type
    lines[0] = json.dumps(doc, sort_keys=True)
    chunks_path.write_text("\n".join(lines) + "\n", encoding="utf-8", newline="\n")
    ok, _ = ctx.validate_ok(inp, work)
    ctx.trap("wrong-type-chunk-rejected", not ok,
             "wrong-typed chunk field passed validation")
    chunks_path.write_bytes(original)

    # Trap 7: xlsx formula cells stay inert source text (never computed)
    fx_sha = sha256_of_bytes(formula_xlsx)
    ok = False
    detail = "formula xlsx bundle missing"
    fb_path = work / "extracted" / f"{fx_sha}.bundle.json"
    if fx_sha in records and fb_path.exists():
        fb = json.loads(fb_path.read_text(encoding="utf-8"))
        text = "\n".join(p["text"] for p in fb["pages"])
        ok = ('=HYPERLINK("http://evil.example/x","click")' in text
              and "legit value" in text and fb["channel"] == "native-text")
        detail = "formula was evaluated, altered, or lost during extraction"
    ctx.trap("xlsx-formula-inert-source-text", ok, detail)

    # Trap 8: csv formula-shaped cells survive verbatim (no interpretation)
    fc_sha = sha256_of_bytes(formula_csv)
    ok = False
    detail = "formula csv bundle missing"
    cb_path = work / "extracted" / f"{fc_sha}.bundle.json"
    if fc_sha in records and cb_path.exists():
        cb = json.loads(cb_path.read_text(encoding="utf-8"))
        text = "".join(p["text"] for p in cb["pages"])
        ok = text == formula_csv.decode("utf-8") and cb["channel"] == "native-text"
        detail = "csv text not preserved verbatim"
    ctx.trap("csv-formula-cells-verbatim", ok, detail)

    # Trap 9: .html extension over PDF magic -> TAMPER_MAGIC_MISMATCH drop
    ok = drops.get("masq.html") == REASON_MAGIC_MISMATCH
    ctx.trap("html-extension-binary-magic-dropped", ok,
             f"got {drops.get('masq.html')}")

    # Trap 10: malformed html accepted, extracted deterministically, inert
    mh_sha = sha256_of_bytes(malformed_html)
    ok = False
    detail = "malformed html bundle missing"
    mh_path = work / "extracted" / f"{mh_sha}.bundle.json"
    if mh_sha in records and mh_path.exists():
        mh = json.loads(mh_path.read_text(encoding="utf-8"))
        vault_ok = (work / records[mh_sha]["vault_relpath"]).read_bytes() == malformed_html
        ok = vault_ok and mh["channel"] == "native-text" and mh["pages"]
        detail = "malformed html crashed extraction or altered vault bytes"
    ctx.trap("malformed-html-never-crashes", bool(ok), detail)

    # Controls
    ctx.control("control-doc-accepted", sha256_of_bytes(control) in records)
    ok, detail = ctx.validate_ok(inp, work)
    ctx.control("restored-corpus-validates", ok, detail)
